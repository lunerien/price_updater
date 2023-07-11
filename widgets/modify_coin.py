from kivy.uix.popup import Popup
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.boxlayout import BoxLayout
from kivy.utils import get_color_from_hex
from openpyxl import load_workbook

from widgets.menu import UNPRESSED_COLOR, PRESSED_COLOR
from lib.coin import Coin
from lib.update import Update
from lib.language import language, Text

DELETE_COLOR = get_color_from_hex("#FF0101e6")
ERROR_COLOR = get_color_from_hex("##c91010F6")


class ModifyCoin(BoxLayout):
    def __init__(self, scrollapp, popup:Popup, coin:Coin):
        super(ModifyCoin, self).__init__()
        self.scrollapp = scrollapp
        self.popup = popup
        self.coin = coin
        self.orientation = "vertical"
        self.opacity = 0.8
        self.coin_name_input = TextInput(text=self.coin.name, size_hint=(1, 0.5))
        self.workbook_name_input = TextInput(text=self.coin.worksheet, size_hint=(1, 0.5))
        self.cell_input = TextInput(text=self.coin.cell, size_hint=(1, 0.5))
        self.add_widget(self.coin_name_input)
        self.add_widget(self.workbook_name_input)
        self.add_widget(self.cell_input)
        buttons = BoxLayout(orientation='horizontal')
        self.add_widget(buttons)
        buttons.add_widget(Button(text=language.get_text(Text.MODIFY.value), on_release=self.modify, size_hint=(0.4, 0.7),
                               background_color=UNPRESSED_COLOR))
        buttons.add_widget(Button(text=language.get_text(Text.DELETE.value), on_release=self.delete, size_hint=(0.4, 0.7),
                               background_color=DELETE_COLOR))
        
    def modify(self, dt):
        dt.background_color=PRESSED_COLOR
        
        workbook = load_workbook(language.read_file()['path_to_xlsx'])
        data = workbook['data']

        test_price = Update().get_token_price(self.coin_name_input.text)
        if workbook != None:
            if test_price != None:
                data.cell(row=1, column=self.coin.id).value = self.coin_name_input.text
                data.cell(row=2, column=self.coin.id).value = self.workbook_name_input.text
                data.cell(row=3, column=self.coin.id).value = self.cell_input.text
                workbook.save(language.read_file()['path_to_xlsx'])

                self.scrollapp.initialize_coins()
                self.popup.dismiss()
            else:
                self.coin_name_input.foreground_color = ERROR_COLOR
        else:
            print("brak arkusza do zapisania!")

    def delete(self, dt):
        dt.background_color=PRESSED_COLOR

        workbook = load_workbook(language.read_file()['path_to_xlsx'])
        data = workbook['data']
        data.cell(row=1, column=self.coin.id).value = "-"
        data.cell(row=2, column=self.coin.id).value = ""
        data.cell(row=3, column=self.coin.id).value = ""
        workbook.save(language.read_file()['path_to_xlsx'])

        self.scrollapp.initialize_coins()
        self.scrollapp.coins.height = self.scrollapp.SPACING + self.scrollapp.COIN_HEIGHT * len(self.scrollapp.coins_tab)
        self.popup.dismiss()
