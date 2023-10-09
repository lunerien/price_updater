from kivymd.uix.scrollview import MDScrollView
from kivymd.uix.gridlayout import MDGridLayout
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from typing import List, Any
from openpyxl import load_workbook
from kivy.clock import Clock

from widgets.coin_button import CoinButton
from lib.asset import Asset
from lib.language import language, Text
from lib.currency import currency, Currency
from lib.update import Update
from lib.config import *


class ScrollApp(MDScrollView):
    SPACING = 2
    COIN_HEIGHT = 40

    def __init__(self) -> None:
        super().__init__()
        self.coins_tab: List[Asset] = list()
        self.coins = MDGridLayout(cols=1, spacing=self.SPACING, size_hint_y=None)
        self.empty_list = Label(
            text=language.get_text(Text.EMPTY_LIST_TEXT.value),
            font_name=font_config,
            font_size=21,
            color=COLOR_ORANGE_THEME,
        )
        self.loading_list = Label(
            text=language.get_text(Text.LOADING_LIST_TEXT.value),
            font_name=font_config,
            font_size=21,
            color=COLOR_ORANGE_THEME,
        )
        self.add_widget(self.loading_list)
        self.coins.height = self.SPACING + self.COIN_HEIGHT * len(self.coins_tab)
        self.bar_color = COLOR_ORANGE_THEME
        self.bar_width = 5
        self.fetch_error: bool = False
        Clock.schedule_interval(self.show_coins, 3)

    def show_coins(self, dt: Any) -> None:
        Clock.unschedule(self.show_coins)
        currency.usd_pln = currency.get_currency(Currency.USD)
        currency.eur_pln = currency.get_currency(Currency.EUR)
        currency.gbp_pln = currency.get_currency(Currency.GBP)
        self.coins_tab = self.get_coins_from_xlsx()
        self.initialize_coins()
        self.clear_widgets()
        self.add_widget(self.coins)
        self.fetch_error_msg()

    def initialize_coins(self) -> None:
        self.coins.height = ScrollApp.SPACING + ScrollApp.COIN_HEIGHT * len(
            self.coins_tab
        )
        self.coins.clear_widgets()
        i = 1
        if len(self.coins_tab):
            self.coins.add_widget(BoxLayout(size_hint=(1, 0.005)))
            for coin in self.coins_tab:
                coin_button = CoinButton(scrollapp=self, coin=coin)
                coin_button.text_color = color_button
                coin_button.md_bg_color = COLOR_ASSET_BUTTON
                i += 1
                if coin.price_eur == "0,0":
                    coin_button.text_color = color_error
                self.coins.add_widget(coin_button)
        else:
            self.coins.add_widget(self.empty_list)
            self.coins.height = self.SPACING + self.COIN_HEIGHT * len(self.coins_tab)

    def get_coins_from_xlsx(self) -> list[Asset]:
        try:
            workbook = load_workbook(language.read_file()["path_to_xlsx"])
        except:
            return []

        if "data" in workbook.sheetnames:
            None
        else:
            workbook.create_sheet("data")
            hidden = workbook["data"]
            hidden.sheet_state = "hidden"
            workbook.save(language.read_file()["path_to_xlsx"])
        data = workbook["data"]

        coins: List[Asset] = []

        i = 1
        while data.cell(row=1, column=i).value != None:
            if data.cell(row=1, column=i).value != "-":
                ticker = data.cell(row=1, column=i).value
                worksheet = data.cell(row=2, column=i).value
                cell = data.cell(row=3, column=i).value
                currency = data.cell(row=4, column=i).value
                price = Update().get_asset_price(ticker)
                try:
                    if price[Currency.PLN] != "0,0":
                        coins.append(
                            Asset(
                                asset_id=i,
                                name=ticker,
                                worksheet=worksheet,
                                cell=cell,
                                price=price,
                                currency=Currency(currency),
                            )
                        )
                    else:
                        self.fetch_error = True
                        coins.append(
                            Asset(
                                asset_id=i,
                                name=ticker,
                                worksheet=worksheet,
                                cell=cell,
                                price={
                                    Currency.USD: "0,0",
                                    Currency.PLN: "0,0",
                                    Currency.EUR: "0,0",
                                    Currency.GBP: "0,0",
                                    Currency.LOGO: price[Currency.LOGO],
                                },
                                currency=Currency(currency),
                            )
                        )
                except:
                    self.fetch_error = True
                    coins.append(
                        Asset(
                            asset_id=i,
                            name=ticker,
                            worksheet=worksheet,
                            cell=cell,
                            price={
                                Currency.USD: "0,0",
                                Currency.PLN: "0,0",
                                Currency.EUR: "0,0",
                                Currency.GBP: "0,0",
                                Currency.LOGO: price[Currency.LOGO],
                            },
                            currency=Currency(currency),
                        )
                    )
            i += 1
        return coins

    def fetch_error_msg(self) -> None:
        if self.fetch_error:
            warning_msg = Popup(
                title_color=COLOR_ORANGE_THEME,
                overlay_color=COLOR_BEHIND_WINDOW,
                separator_color=COLOR_ORANGE_THEME,
                size_hint=(None, None),
                size=(350, 200),
                auto_dismiss=True,
                title_font=font_config,
                title=language.get_text(Text.FETCH_ERROR_TITLE.value),
                background_color=COLOR_WINDOW,
            )
            warning_content = Label(
                text=language.get_text(Text.CONNECTION_LOST.value)
                if currency.connection_lost
                else language.get_text(Text.FETCH_ERROR_MSG.value),
                font_name=font_config,
                color=color_error,
            )
            warning_msg.content = warning_content
            warning_msg.open(animation=True)
