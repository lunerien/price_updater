from typing import Tuple, List, Any
from time import sleep
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils.exceptions import InvalidFileException
from bs4 import BeautifulSoup
from requests import get, exceptions, Response

from lib.language import language
from lib.currency import currency, Currency
from lib.asset import Asset


class Update:
    def __init__(self) -> None:
        self.dec: int = 3
        self.api_status: str = language.get_api_status()

    def update(self, coins: List[Asset]) -> bool:
        try:
            workbook: Workbook = self.try_load_workbook()
            if workbook is not None:
                data = workbook["data"]
                i: int = 1
                while data.cell(row=1, column=i).value is not None:
                    if data.cell(row=1, column=i).value != "-":
                        price = next(
                            self.get_price(coin)
                            for coin in coins
                            if coin.name == data.cell(row=1, column=i).value
                        )
                        sheet = workbook[data.cell(row=2, column=i).value]
                        if price != "0,0":
                            sheet[data.cell(row=3, column=i).value] = price
                    i += 1
                workbook.save(language.read_file()["path_to_xlsx"])
                return True
            return False
        except UnboundLocalError:
            return False

    def get_price(self, coin: Asset) -> str:
        match coin.chosen_currency:
            case Currency.USD:
                return coin.price_usd
            case Currency.PLN:
                return coin.price_pln
            case Currency.EUR:
                return coin.price_eur
            case Currency.GBP:
                return coin.price_gbp
        return "0,0"

    def get_asset_price(self, ticker: str) -> dict[Currency, str]:
        fiat_assets: Tuple[str, ...] = ("eur", "gbp", "usd")
        metal_assets: Tuple[str, ...] = ("xau", "xag")
        etf_assets: Tuple[str, ...] = ("swda-etf", "emim-etf")

        if ticker in fiat_assets:
            return self.get_fiat_price(ticker)
        if ticker in metal_assets:
            return self.get_metal_price(ticker)
        if ticker in etf_assets:
            return self.get_etf_price(ticker)
        return self.get_crypto_price(ticker)

    def try_load_workbook(self) -> Workbook | None:
        try:
            workbook: Workbook = load_workbook(language.read_file()["path_to_xlsx"])
            return workbook
        except InvalidFileException:
            print("we need xlsx file!")
        except KeyError:
            print("please check xlsx format file!")
        except FileNotFoundError:
            print("file missing :D")
        return None

    def _exception_catch(self, error: Any) -> float:
        if isinstance(error, (exceptions.ConnectionError, exceptions.ReadTimeout)):
            currency.connection_lost = True
        return 0.0

    def get_fiat_price(self, ticker: str) -> dict[Currency, str]:
        try:
            match (ticker):
                case "usd":
                    return {
                        Currency.USD: "1,0",
                        Currency.PLN: str(currency.usd_pln).replace(".", ","),
                        Currency.EUR: str(
                            round(currency.usd_pln / currency.eur_pln, self.dec)
                        ).replace(".", ","),
                        Currency.GBP: str(
                            round(currency.usd_pln / currency.gbp_pln, self.dec)
                        ).replace(".", ","),
                        Currency.LOGO: f"./images/asset_logo/{ticker}.png",
                    }
                case "gbp":
                    return {
                        Currency.USD: str(
                            round(currency.gbp_pln / currency.usd_pln, self.dec)
                        ).replace(".", ","),
                        Currency.PLN: str(currency.gbp_pln).replace(".", ","),
                        Currency.EUR: str(
                            round(currency.gbp_pln / currency.eur_pln, self.dec)
                        ).replace(".", ","),
                        Currency.GBP: "1,0",
                        Currency.LOGO: f"./images/asset_logo/{ticker}.png",
                    }
                case "eur":
                    return {
                        Currency.USD: str(
                            round(currency.eur_pln / currency.usd_pln, self.dec)
                        ).replace(".", ","),
                        Currency.PLN: str(currency.eur_pln).replace(".", ","),
                        Currency.EUR: "1,0",
                        Currency.GBP: str(
                            round(currency.eur_pln / currency.gbp_pln, self.dec)
                        ).replace(".", ","),
                        Currency.LOGO: f"./images/asset_logo/{ticker}.png",
                    }
                case _:
                    return {
                        Currency.USD: "0,0",
                        Currency.PLN: "0,0",
                        Currency.EUR: "0,0",
                        Currency.GBP: "0,0",
                        Currency.LOGO: f"./images/asset_logo/{ticker}.png",
                    }
        except ZeroDivisionError:
            return {
                Currency.USD: "0,0",
                Currency.PLN: "0,0",
                Currency.EUR: "0,0",
                Currency.GBP: "0,0",
                Currency.LOGO: f"./images/asset_logo/{ticker}.png",
            }

    def get_metal_price(self, ticker: str) -> dict[Currency, str]:
        def get_price() -> float:
            try:
                url = f"https://www.cnbc.com/quotes/{ticker}="
                page = get(url, timeout=7)
                page_content = BeautifulSoup(page.content, "html.parser")
                onpage = str(page_content.find("span", class_="QuoteStrip-lastPrice"))
                onpage = onpage.replace('<span class="QuoteStrip-lastPrice">', "")
                onpage = onpage.replace("</span>", "")
                price = onpage.replace(",", "")
                return float(price)
            except (
                ValueError,
                ZeroDivisionError,
                TypeError,
                exceptions.ConnectionError,
                exceptions.ReadTimeout,
            ) as error:
                value = self._exception_catch(error)
                return value

        price_exact: float = get_price()
        try:
            return {
                Currency.USD: str(price_exact).replace(".", ","),
                Currency.PLN: str(
                    round(price_exact * currency.usd_pln, self.dec)
                ).replace(".", ","),
                Currency.EUR: str(
                    round(price_exact * (currency.usd_pln / currency.eur_pln), self.dec)
                ).replace(".", ","),
                Currency.GBP: str(
                    round(price_exact * (currency.usd_pln / currency.gbp_pln), self.dec)
                ).replace(".", ","),
                Currency.LOGO: f"./images/asset_logo/{ticker}.png",
            }
        except ZeroDivisionError:
            return {
                Currency.USD: "0,0",
                Currency.PLN: "0,0",
                Currency.EUR: "0,0",
                Currency.GBP: "0,0",
                Currency.LOGO: f"./images/asset_logo/{ticker}.png",
            }

    def get_etf_price(self, ticker: str) -> dict[Currency, str]:
        link: str
        if ticker == "swda-etf":
            link = "https://markets.ft.com/data/etfs/tearsheet/summary?s=SWDA:LSE:GBX"
        else:
            link = "https://markets.ft.com/data/etfs/tearsheet/summary?s=EMIM:LSE:GBX"

        def get_price() -> float:
            try:
                page = get(link, timeout=7)
                page_content = BeautifulSoup(page.content, "html.parser")
                for onpage in page_content.find(
                    "span", class_="mod-ui-data-list__value"
                ):
                    page_str = str(onpage)
                    page_str = page_str.replace(",", "")
                    price = page_str[0:4]
                    return float(price)
            except (
                ValueError,
                ZeroDivisionError,
                TypeError,
                exceptions.ConnectionError,
                exceptions.ReadTimeout,
            ) as error:
                value = self._exception_catch(error)
                return value
            return 0.0

        price_exact = get_price()
        try:
            return {
                Currency.USD: str(
                    round(price_exact * (currency.gbp_pln / currency.usd_pln), self.dec)
                ).replace(".", ","),
                Currency.PLN: str(
                    round(price_exact * currency.gbp_pln, self.dec)
                ).replace(".", ","),
                Currency.EUR: str(
                    round(price_exact * (currency.gbp_pln / currency.eur_pln), self.dec)
                ).replace(".", ","),
                Currency.GBP: str(price_exact).replace(".", ","),
                Currency.LOGO: "./images/asset_logo/ishares.png",
            }
        except ZeroDivisionError:
            return {
                Currency.USD: "0,0",
                Currency.PLN: "0,0",
                Currency.EUR: "0,0",
                Currency.GBP: "0,0",
                Currency.LOGO: "./images/asset_logo/ishares.png",
            }

    def get_crypto_price(self, ticker: str) -> dict[Currency, str]:
        if ticker[-4:] == "_api":
            ticker = ticker.replace(ticker[-4:], "")
        url: str = f"https://coinmarketcap.com/currencies/{ticker.lower()}"
        page: Response = get(url, timeout=7)
        page_content = BeautifulSoup(page.content, "html.parser")

        price_float: float = 0.0

        def get_logo() -> str:
            try:
                urla: str = f"https://coinmarketcap.com/currencies/{ticker}"
                pagea: Response = get(urla, timeout=5)
                page_contenta = BeautifulSoup(pagea.content, "html.parser")
                dataa = page_contenta.find("div", class_="sc-65e7f566-0 kYcmYb")
                raw_dataa = dataa.find("img", src=True)
                return raw_dataa["src"]
            except (exceptions.ConnectionError, AttributeError):
                pass
            return "./images/asset_logo/none.png"

        def get_price_from_api() -> float:
            try:
                url_api: str = "https://api.coingecko.com/api/v3/simple/price"
                params: dict[str, str] = {"ids": ticker, "vs_currencies": "USD"}

                response: Response = get(url_api, params=params, timeout=5)
                if response.status_code == 200:
                    data = response.json()
                    price = data[ticker]["usd"]
                    return float(price)
                return 0.0
            except (
                AttributeError,
                ValueError,
                ZeroDivisionError,
                TypeError,
                KeyError,
                exceptions.ConnectionError,
                exceptions.ReadTimeout,
            ) as error:
                self._exception_catch(error)
                return 0.0

        def get_price() -> float:
            try:
                web = str(
                    page_content.find("span", class_="sc-65e7f566-0 clvjgF base-text")
                )
                price_str = web.replace(
                    '<span class="sc-65e7f566-0 clvjgF base-text" data-test="text-cdp-price-display">$',
                    "",
                )
                price_str = price_str.replace("</span>", "")
                price_str = price_str.replace(",", "")
                price = float(price_str)
                return price
            except (
                AttributeError,
                ValueError,
                ZeroDivisionError,
                TypeError,
                KeyError,
                exceptions.ConnectionError,
                exceptions.ReadTimeout,
            ) as error:
                self._exception_catch(error)
                return 0.0

        if self.api_status == "api-off":
            price_float = get_price()
            if price_float == 0.0:
                price_float = get_price_from_api()
        else:
            i = 0
            while i < 4 and price_float == 0.0:
                sleep(0.2)
                price_float = get_price_from_api()
                i += 1
            if price_float == 0.0:
                price_float = get_price()
        ######################################################
        price_usd = (
            str(round(price_float, 6))
            if str(price_float)[:2] == "0."
            else str(round(price_float, 2))
        )
        price_usd = price_usd.replace(".", ",")
        ######################################################
        price_pln_float = price_float * currency.return_price(Currency.USD)
        price_pln = (
            str(round(price_pln_float, 6))
            if str(price_float)[:2] == "0."
            else str(round(price_pln_float, 2))
        )
        price_pln = price_pln.replace(".", ",")
        ######################################################
        price_eur_float = price_float * (
            currency.return_price(Currency.USD) / currency.return_price(Currency.EUR)
        )
        price_eur = (
            str(round(price_eur_float, 6))
            if str(price_float)[:2] == "0."
            else str(round(price_eur_float, 2))
        )
        price_eur = price_eur.replace(".", ",")
        ######################################################
        price_gbp_float = price_float * (
            currency.return_price(Currency.USD) / currency.return_price(Currency.GBP)
        )
        price_gbp = (
            str(round(price_gbp_float, 6))
            if str(price_float)[:2] == "0."
            else str(round(price_gbp_float, 2))
        )
        price_gbp = price_gbp.replace(".", ",")
        ######################################################
        http_logo: str = get_logo()

        return {
            Currency.USD: price_usd,
            Currency.PLN: price_pln,
            Currency.EUR: price_eur,
            Currency.GBP: price_gbp,
            Currency.LOGO: http_logo,
        }
